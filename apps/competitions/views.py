from django.contrib.messages.views import SuccessMessageMixin
import logging

from django.urls import reverse_lazy
from django.views.generic import CreateView, UpdateView, ListView, DetailView
from django.contrib.auth.mixins import UserPassesTestMixin, LoginRequiredMixin
from django.utils.translation import gettext_lazy as _
from django.views import View
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.db.models import Count, Q, Avg, F, Case, When, Value, CharField
from django.utils import timezone
from datetime import timedelta
from django.views.generic import TemplateView
from django.http import HttpResponse
import openpyxl

from apps.users.models import User
from apps.plays.models import Play
from apps.reviews.models import Review
from .models import Competition
from .mixins import CompetitionContextMixin
from .forms import CompetitionCreationForm, CompetitionChangeForm
from .services import sync_plays_from_google_sheet
from apps.reviews.services import auto_assign_phase2

logger = logging.getLogger(__name__)


class CompetitionCreateView(
    LoginRequiredMixin, UserPassesTestMixin, SuccessMessageMixin, CreateView
):
    model = Competition
    template_name = "create_update.html"
    form_class = CompetitionCreationForm
    success_message = _("✅ %(title)s was created successfully")

    def get_success_url(self):
        return self.object.get_absolute_url()

    def test_func(self):
        return self.request.user.is_superuser


class CompetitionUpdateView(
    LoginRequiredMixin, UserPassesTestMixin, SuccessMessageMixin, UpdateView
):
    model = Competition
    template_name = "create_update.html"
    form_class = CompetitionChangeForm
    success_message = _("✅ %(title)s was updated successfully")

    def form_valid(self, form):
        old_status = self.get_object().status
        response = super().form_valid(form)

        if (
            old_status != Competition.Status.PHASE_2
            and self.object.status == Competition.Status.PHASE_2
        ):
            count = auto_assign_phase2(self.object)
            if count > 0:
                messages.success(
                    self.request,
                    _("✅ %(count)s Phase 2 reviews have been assigned.")
                    % {"count": count},
                )

        return response

    def get_success_url(self):
        return self.object.get_absolute_url()

    def test_func(self):
        user = self.request.user
        if user.is_superuser:
            return True

        return user.competition_roles.filter(
            competition=self.get_object(), role__in=["admin"], is_active=True
        ).exists()


class CompetitionListView(LoginRequiredMixin, ListView):
    model = Competition
    template_name = "competition_list.html"
    context_object_name = "competitions"

    def get_queryset(self):
        user = self.request.user
        if user.is_superuser:
            return Competition.objects.all()

        return Competition.objects.filter(
            roles__user=user, roles__is_active=True
        ).distinct()


class CompetitionDetailView(
    LoginRequiredMixin, UserPassesTestMixin, CompetitionContextMixin, DetailView
):
    model = Competition
    template_name = "competition_detail.html"
    context_object_name = "competition"

    def test_func(self):
        user = self.request.user
        if user.is_superuser:
            return True

        return user.competition_roles.filter(competition=self.get_object()).exists()


class CompetitionSyncView(LoginRequiredMixin, UserPassesTestMixin, View):
    def post(self, request, *args, **kwargs):
        competition_slug = self.kwargs.get("competition_slug")
        competition = get_object_or_404(Competition, slug=competition_slug)

        try:
            count = sync_plays_from_google_sheet(competition)
            messages.success(
                request, _("Successfully synced %(count)s plays.") % {"count": count}
            )
        except Exception as e:
            messages.error(
                request,
                _("Sync error: %(error)s") % {"error": f"{type(e).__name__}: {e}"},
            )
            logger.exception("Sync error")

        return redirect("plays:list", competition_slug=competition.slug)

    def test_func(self):
        return (
            self.request.user.is_superuser
            or self.request.user.competition_roles.filter(
                competition__slug=self.kwargs.get("competition_slug"),
                role__in=["admin", "moderator"],
                is_active=True,
            ).exists()
        )


class CompetitionAnalyticsView(
    LoginRequiredMixin, UserPassesTestMixin, CompetitionContextMixin, TemplateView
):
    template_name = "analytics_dashboard.html"

    def test_func(self):
        competition = self.get_competition()
        if self.request.user.is_superuser:
            return True
        return self.request.user.get_role(competition) in ["admin", "moderator"]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        competition = self.get_competition()

        default_phase = competition.status
        if default_phase not in ["phase_1", "phase_2"]:
            default_phase = "all"

        selected_phase = self.request.GET.get("phase", default_phase)

        plays_qs = Play.objects.filter(competition=competition, is_active=True)

        users_qs = User.objects.filter(
            competition_roles__competition=competition,
            competition_roles__role="reader",
            competition_roles__is_active=True,
        ).distinct()

        reader_count = users_qs.count()

        review_filters = Q(play__competition=competition, is_obsolete=False)
        if selected_phase in ["phase_1", "phase_2"]:
            review_filters &= Q(phase=selected_phase)

        reviews_qs = Review.objects.filter(review_filters)

        user_review_filters = Q(
            review__play__competition=competition, review__is_obsolete=False
        )
        if selected_phase in ["phase_1", "phase_2"]:
            user_review_filters &= Q(review__phase=selected_phase)

        top_readers_qs = users_qs.annotate(
            submitted_reviews=Count(
                "review",
                filter=user_review_filters & Q(review__status=Review.Status.SUBMITTED),
            ),
            submitted_yes=Count(
                "review",
                filter=user_review_filters
                & Q(review__status=Review.Status.SUBMITTED, review__verdict=True),
            ),
            submitted_no=Count(
                "review",
                filter=user_review_filters
                & Q(review__status=Review.Status.SUBMITTED, review__verdict=False),
            ),
            active_reviews=Count(
                "review",
                filter=user_review_filters
                & Q(review__status__in=[Review.Status.ASSIGNED, Review.Status.DRAFT]),
            ),
            avg_reading_time=Avg(
                F("review__submitted_at") - F("review__created_at"),
                filter=user_review_filters & Q(review__status=Review.Status.SUBMITTED),
            ),
        ).order_by("-submitted_reviews")

        top_readers = list(top_readers_qs)
        for reader in top_readers:
            if reader.avg_reading_time:
                days = reader.avg_reading_time.days
                hours = reader.avg_reading_time.seconds // 3600
                reader.speed_str = f"{days}d {hours}h" if days > 0 else f"{hours}h"
            else:
                reader.speed_str = "-"

        p1_yes = Count(
            "reviews",
            filter=Q(
                reviews__phase="phase_1",
                reviews__status=Review.Status.SUBMITTED,
                reviews__verdict=True,
                reviews__is_obsolete=False,
            ),
        )
        p1_no = Count(
            "reviews",
            filter=Q(
                reviews__phase="phase_1",
                reviews__status=Review.Status.SUBMITTED,
                reviews__verdict=False,
                reviews__is_obsolete=False,
            ),
        )
        p2_yes = Count(
            "reviews",
            filter=Q(
                reviews__phase="phase_2",
                reviews__status=Review.Status.SUBMITTED,
                reviews__verdict=True,
                reviews__is_obsolete=False,
            ),
        )
        p2_no = Count(
            "reviews",
            filter=Q(
                reviews__phase="phase_2",
                reviews__status=Review.Status.SUBMITTED,
                reviews__verdict=False,
                reviews__is_obsolete=False,
            ),
        )

        play_review_filters = Q(
            reviews__is_obsolete=False, reviews__status=Review.Status.SUBMITTED
        )
        if selected_phase in ["phase_1", "phase_2"]:
            play_review_filters &= Q(reviews__phase=selected_phase)

        plays_overview = (
            plays_qs.annotate(
                total_submitted=Count("reviews", filter=play_review_filters),
                phase_1_yes=p1_yes,
                phase_1_no=p1_no,
                phase_2_yes=p2_yes,
                phase_2_no=p2_no,
            )
            .annotate(
                current_status=Case(
                    When(
                        phase_1_no__gte=2,
                        then=Value(str(_("Eliminated in Phase 1 ❌"))),
                    ),
                    When(phase_1_yes__gte=2, then=Value(str(_("Phase 2")))),
                    default=Value(str(_("Phase 1"))),
                    output_field=CharField(),
                )
            )
            .order_by("-total_submitted")
        )

        qualifying_plays_count = plays_overview.filter(phase_1_yes__gte=2).count()

        if selected_phase == "phase_1":
            plays_overview = plays_overview.filter(phase_1_yes__lt=2)
        elif selected_phase == "phase_2":
            plays_overview = plays_overview.filter(phase_1_yes__gte=2)

        yes_filters = play_review_filters & Q(reviews__verdict=True)
        no_filters = play_review_filters & Q(reviews__verdict=False)

        controversial_plays = (
            plays_qs.annotate(
                yes_votes=Count("reviews", filter=yes_filters),
                no_votes=Count("reviews", filter=no_filters),
            )
            .filter(yes_votes__gt=0, no_votes__gt=0)
            .order_by("-yes_votes", "-no_votes")
        )

        total_yes = reviews_qs.filter(
            status=Review.Status.SUBMITTED, verdict=True
        ).count()
        total_no = reviews_qs.filter(
            status=Review.Status.SUBMITTED, verdict=False
        ).count()

        pending_actions = None
        if selected_phase == "phase_1":
            pending_actions = (
                Review.objects.filter(
                    play__competition=competition,
                    phase=Review.Phase.PHASE_1,
                    status__in=[Review.Status.ASSIGNED, Review.Status.DRAFT],
                    is_obsolete=False,
                )
                .select_related("play", "reader")
                .order_by("created_at")[:10]
            )
        elif selected_phase == "phase_2":
            pending_actions = sorted(top_readers, key=lambda x: x.submitted_reviews)[
                :10
            ]

        eta_days = None
        total_done = reviews_qs.filter(status=Review.Status.SUBMITTED).count()
        total_target = 0

        seven_days_ago = timezone.now() - timedelta(days=7)
        recent_submissions = reviews_qs.filter(
            status=Review.Status.SUBMITTED, submitted_at__gte=seven_days_ago
        ).count()
        velocity_per_day = recent_submissions / 7.0

        active_plays_count = plays_qs.count()
        if selected_phase == "phase_1":
            total_target = active_plays_count * 2
        elif selected_phase == "phase_2":
            total_target = qualifying_plays_count * reader_count
        else:
            total_target = (active_plays_count * 2) + (
                qualifying_plays_count * reader_count
            )

        remaining_tasks = max(0, total_target - total_done)
        if velocity_per_day > 0 and remaining_tasks > 0:
            eta_days = round(remaining_tasks / velocity_per_day)

        context.update(
            {
                "selected_phase": selected_phase,
                "top_readers": top_readers,
                "plays_overview": plays_overview,
                "controversial_plays": controversial_plays,
                "total_yes": total_yes,
                "total_no": total_no,
                "pending_actions": pending_actions,
                "eta_days": eta_days,
                "velocity_per_day": round(velocity_per_day, 1),
                "progress_percent": (
                    int((total_done / total_target * 100)) if total_target > 0 else 0
                ),
            }
        )

        return context


class CompetitionExportExcelView(
    LoginRequiredMixin, UserPassesTestMixin, CompetitionContextMixin, View
):
    def get(self, request, *args, **kwargs):
        competition = self.get_competition()

        wb = openpyxl.Workbook()

        ws_plays = wb.active
        ws_plays.title = "Plays"
        ws_plays.append(
            [
                "ID",
                "Title",
                "Author Email",
                "Author First Name",
                "Author Last Name",
                "Status",
                "Total Reviews",
                "Phase 1 Yes",
                "Phase 1 No",
                "Phase 2 Yes",
                "Phase 2 No",
            ]
        )

        plays = Play.objects.filter(competition=competition).prefetch_related("reviews")
        for play in plays:
            reviews = play.reviews.filter(is_obsolete=False, status="submitted")
            ws_plays.append(
                [
                    play.id,
                    play.title,
                    play.author_email,
                    play.author_first_name,
                    play.author_last_name or "",
                    "Active" if play.is_active else "Inactive",
                    reviews.count(),
                    reviews.filter(phase="phase_1", verdict=True).count(),
                    reviews.filter(phase="phase_1", verdict=False).count(),
                    reviews.filter(phase="phase_2", verdict=True).count(),
                    reviews.filter(phase="phase_2", verdict=False).count(),
                ]
            )

        ws_reviews = wb.create_sheet(title="Reviews")
        ws_reviews.append(
            [
                "Play ID",
                "Play Title",
                "Reader Username",
                "Phase",
                "Verdict",
                "Comment",
                "Submitted At",
            ]
        )

        all_reviews = Review.objects.filter(
            play__competition=competition, status="submitted", is_obsolete=False
        ).select_related("play", "reader")

        for review in all_reviews:
            ws_reviews.append(
                [
                    review.play.id,
                    review.play.title,
                    review.reader.username,
                    review.get_phase_display(),
                    "Yes" if review.verdict else "No",
                    review.comment,
                    (
                        review.submitted_at.strftime("%Y-%m-%d %H:%M")
                        if review.submitted_at
                        else ""
                    ),
                ]
            )

        ws_readers = wb.create_sheet(title="Readers")
        ws_readers.append(
            ["Username", "Telegram", "Role", "Submitted (Yes)", "Submitted (No)"]
        )

        users = (
            User.objects.filter(
                competition_roles__competition=competition,
                competition_roles__is_active=True,
            )
            .annotate(
                yes_votes=Count(
                    "review",
                    filter=Q(
                        review__play__competition=competition,
                        review__status="submitted",
                        review__verdict=True,
                    ),
                ),
                no_votes=Count(
                    "review",
                    filter=Q(
                        review__play__competition=competition,
                        review__status="submitted",
                        review__verdict=False,
                    ),
                ),
            )
            .distinct()
        )

        for u in users:
            role = u.get_role(competition)
            ws_readers.append(
                [
                    u.username,
                    u.telegram_username or "",
                    role,
                    u.yes_votes,
                    u.no_votes,
                ]
            )

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{competition.slug}_full_export.xlsx"'
        )
        wb.save(response)

        return response

    def test_func(self):
        competition = self.get_competition()
        if self.request.user.is_superuser:
            return True
        return self.request.user.get_role(competition) in ["admin", "moderator"]
